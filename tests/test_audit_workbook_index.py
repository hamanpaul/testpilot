from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from testpilot.audit.workbook_index import (
    WorkbookRow,
    build_index,
    normalize_api,
    normalize_object,
)

FIXTURE = Path(__file__).resolve().parent / "fixtures" / "audit" / "sample_workbook.xlsx"


def test_normalize_object_strips_trailing_dot_and_collapses_index():
    assert normalize_object("WiFi.Radio.{i}.") == "WiFi.Radio.{i}"
    assert normalize_object("WiFi.Radio.1.") == "WiFi.Radio.{i}"
    assert normalize_object("WiFi.Radio.{i}.IEEE80211ax.") == "WiFi.Radio.{i}.IEEE80211ax"
    assert normalize_object("WiFi.1.2.") == "WiFi.{i}.{i}"


def test_normalize_api_strips_whitespace_preserves_case():
    assert normalize_api("  Noise  ") == "Noise"
    assert normalize_api("SRGBSSColorBitmap") == "SRGBSSColorBitmap"
    assert normalize_api("noise") != normalize_api("Noise")


def test_build_index_creates_lookup():
    index = build_index(FIXTURE, sheet_name="Wifi_LLAPI")

    key = (normalize_object("WiFi.Radio.{i}."), normalize_api("Noise"))
    assert key in index

    rows = index[key]
    assert len(rows) == 1
    assert rows[0] == WorkbookRow(
        raw_row_index=2,
        object_path="WiFi.Radio.{i}.",
        api="Noise",
        test_steps="ubus-cli WiFi.Radio.1.Noise?",
        command_output="expected",
        result_5g="Pass",
        result_6g="Pass",
        result_24g="Pass",
    )


def test_build_index_detects_ambiguity():
    index = build_index(FIXTURE, sheet_name="Wifi_LLAPI")
    key = (normalize_object("WiFi.Radio.{i}."), normalize_api("DuplicateApi"))
    assert len(index[key]) == 2
    assert index[key][0].command_output == ""


def test_build_index_missing_key():
    index = build_index(FIXTURE, sheet_name="Wifi_LLAPI")
    key = (normalize_object("WiFi.Foo."), normalize_api("Bar"))
    assert key not in index


def test_build_index_uses_column_letter_overrides(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Custom"
    sheet.append(["API", "CmdOut", "5G", "2.4G", "Object", "6G", "Steps"])
    sheet.append(["FooApi", "stdout", "Pass", "Fail", "WiFi.SSID.1.", "Pass", "step body"])

    workbook_path = tmp_path / "column-overrides.xlsx"
    workbook.save(workbook_path)

    overrides = {
        "api": "A",
        "command_output": "B",
        "result_5g": "C",
        "result_24g": "D",
        "object": "E",
        "result_6g": "F",
        "test_steps": "G",
    }

    index = build_index(workbook_path, sheet_name="Custom", column_overrides=overrides)
    assert (normalize_object("WiFi.SSID.1."), normalize_api("FooApi")) in index


def test_build_index_raises_on_missing_required_columns(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Broken"
    sheet.append(["Object", "API"])
    sheet.append(["WiFi.Radio.1.", "Noise"])

    workbook_path = tmp_path / "broken.xlsx"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="missing required columns"):
        build_index(workbook_path, sheet_name="Broken")


def test_build_index_skips_rows_that_normalize_to_empty_keys(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Whitespace"
    sheet.append(["Object", "API", "Test Steps", "Command Output", "5G", "6G", "2.4G"])
    sheet.append(["   ", "Noise", "ignored", "", "", "", ""])
    sheet.append(["WiFi.Radio.1.", "   ", "ignored", "", "", "", ""])
    sheet.append(["WiFi.Radio.1.", "Noise", "kept", "stdout", "Pass", "Pass", "Pass"])

    workbook_path = tmp_path / "whitespace.xlsx"
    workbook.save(workbook_path)

    index = build_index(workbook_path, sheet_name="Whitespace")
    assert list(index.keys()) == [
        (normalize_object("WiFi.Radio.1."), normalize_api("Noise"))
    ]
