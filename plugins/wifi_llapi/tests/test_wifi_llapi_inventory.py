from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook

from testpilot.reporting.wifi_llapi_inventory import audit_wifi_llapi_inventory


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["A4"] = "WiFi.Radio.{i}."
    ws["C4"] = "getRadioStats()"
    ws["A5"] = "WiFi.SSID.{i}."
    ws["C5"] = "getSSIDStats()"
    ws["A6"] = "WiFi.AccessPoint.{i}."
    ws["C6"] = "getAPStats()"
    ws["A7"] = "WiFi.AccessPoint.{i}.AssociatedDevice.{i}."
    ws["C7"] = "AssociatedDeviceNumberOfEntries"
    wb.save(path)
    wb.close()


def _write_case(path: Path, *, case_id: str, name: str, row: int, obj: str, api: str) -> None:
    payload = {
        "id": case_id,
        "name": name,
        "topology": {"devices": {"DUT": {}, "STA": {}}},
        "steps": [
            {
                "id": "step-1",
                "action": "noop",
                "target": "DUT",
            }
        ],
        "pass_criteria": [{"field": "status", "operator": "equals", "expected": "Pass"}],
        "source": {
            "row": row,
            "object": obj,
            "api": api,
        },
    }
    path.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")


def test_audit_classifies_missing_drifted_duplicate_and_extra_cases(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _write_template(template)

    _write_case(
        cases_dir / "D004_canonical.yaml",
        case_id="wifi-llapi-D004-canonical",
        name="getRadioStats()",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )
    _write_case(
        cases_dir / "D009_stale.yaml",
        case_id="wifi-llapi-D009-stale",
        name="getSSIDStats()",
        row=9,
        obj="WiFi.SSID.{i}.",
        api="getSSIDStats()",
    )
    _write_case(
        cases_dir / "D006_canonical.yaml",
        case_id="wifi-llapi-D006-canonical",
        name="getAPStats()",
        row=6,
        obj="WiFi.AccessPoint.{i}.",
        api="getAPStats()",
    )
    _write_case(
        cases_dir / "D006_duplicate.yaml",
        case_id="wifi-llapi-D006-duplicate",
        name="getAPStats()",
        row=6,
        obj="WiFi.AccessPoint.{i}.",
        api="getAPStats()",
    )
    _write_case(
        cases_dir / "D999_extra.yaml",
        case_id="wifi-llapi-D999-extra",
        name="not-in-template()",
        row=999,
        obj="WiFi.NotInTemplate.{i}.",
        api="notInTemplate()",
    )

    audit = audit_wifi_llapi_inventory(template, cases_dir)

    assert audit.missing_rows == (7,)
    assert audit.case_status_counts == {
        "canonical": 2,
        "drifted": 1,
        "duplicate": 1,
        "extra": 1,
    }
    assert audit.rows[4].status == "canonical"
    assert audit.rows[5].status == "drifted"
    assert audit.rows[6].status == "duplicate"
    assert audit.rows[7].status == "missing"
    assert audit.rows[5].canonical_case_file is None
    assert audit.rows[6].canonical_case_file == "D006_canonical.yaml"
    assert audit.cases["D009_stale.yaml"].status == "drifted"
    assert audit.cases["D009_stale.yaml"].resolved_row == 5
    assert audit.cases["D006_duplicate.yaml"].status == "duplicate"
    assert audit.cases["D006_duplicate.yaml"].resolved_row == 6
    assert audit.cases["D999_extra.yaml"].status == "extra"


def test_audit_to_dict_is_machine_checkable(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _write_template(template)
    _write_case(
        cases_dir / "D004_canonical.yaml",
        case_id="wifi-llapi-D004-canonical",
        name="getRadioStats()",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )

    audit = audit_wifi_llapi_inventory(template, cases_dir)
    payload = audit.to_dict()

    assert payload["missing_rows"] == [5, 6, 7]
    assert payload["cases"]["D004_canonical.yaml"]["status"] == "canonical"
    assert payload["rows"]["4"]["canonical_case_file"] == "D004_canonical.yaml"


def test_audit_marks_unresolved_in_template_family_as_drifted(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _write_template(template)
    from openpyxl import load_workbook

    wb = load_workbook(template)
    ws = wb["Wifi_LLAPI"]
    ws["A8"] = "WiFi.SSID.{i}."
    ws["C8"] = "getSSIDStats()"
    wb.save(template)
    wb.close()
    _write_case(
        cases_dir / "D099_family_drift.yaml",
        case_id="wifi-llapi-D099-family-drift",
        name="getSSIDStats()",
        row=99,
        obj="WiFi.SSID.{i}.",
        api="getSSIDStats()",
    )

    audit = audit_wifi_llapi_inventory(template, cases_dir)

    assert audit.cases["D099_family_drift.yaml"].status == "drifted"
    assert audit.cases["D099_family_drift.yaml"].reason == "unresolved_object_api_family"


def test_audit_treats_stale_id_fragment_as_drifted_metadata(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    cases_dir = tmp_path / "cases"
    cases_dir.mkdir()
    _write_template(template)
    _write_case(
        cases_dir / "D004_stale_id.yaml",
        case_id="wifi-llapi-D099-stale-id",
        name="getRadioStats()",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )

    audit = audit_wifi_llapi_inventory(template, cases_dir)

    assert audit.rows[4].status == "drifted"
    assert audit.rows[4].canonical_case_file is None
    assert audit.cases["D004_stale_id.yaml"].status == "drifted"
    assert audit.cases["D004_stale_id.yaml"].resolved_row == 4
