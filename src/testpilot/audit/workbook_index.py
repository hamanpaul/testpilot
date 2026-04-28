"""Workbook xlsx -> (object, api) semantic-key index."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_SPECIFIC_INDEX_RE = re.compile(r"\.(\d+)\.")


@dataclass(frozen=True)
class WorkbookRow:
    raw_row_index: int
    object_path: str
    api: str
    test_steps: str
    command_output: str
    result_5g: str
    result_6g: str
    result_24g: str


def normalize_object(value: str | None) -> str:
    """Normalize source.object for semantic lookup."""
    if value is None:
        return ""
    normalized = str(value).strip()
    normalized = _SPECIFIC_INDEX_RE.sub(".{i}.", normalized)
    while normalized.endswith("."):
        normalized = normalized[:-1]
    return normalized


def normalize_api(value: str | None) -> str:
    """Normalize source.api for semantic lookup."""
    if value is None:
        return ""
    return str(value).strip()


def _column_letter_to_index(letter: str) -> int:
    """Convert Excel column letter to zero-based index."""
    normalized = letter.strip().upper()
    if not normalized:
        raise ValueError("column letter override cannot be empty")

    index = 0
    for ch in normalized:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"invalid column letter: {letter!r}")
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index - 1


def _auto_discover_columns(header_row: list[Any]) -> dict[str, int]:
    headers = [str(cell or "").strip().lower() for cell in header_row]
    columns: dict[str, int] = {}

    for idx, header in enumerate(headers):
        if header == "object":
            columns["object"] = idx
        elif header == "api":
            columns["api"] = idx
        elif header == "test steps":
            columns["test_steps"] = idx
        elif header == "command output":
            columns["command_output"] = idx
        elif header == "5g":
            columns["result_5g"] = idx
        elif header == "6g":
            columns["result_6g"] = idx
        elif header == "2.4g":
            columns["result_24g"] = idx

    required = (
        "object",
        "api",
        "test_steps",
        "command_output",
        "result_5g",
        "result_6g",
        "result_24g",
    )
    missing = [key for key in required if key not in columns]
    if missing:
        raise ValueError(f"workbook headers missing required columns: {missing}")
    return columns


def build_index(
    workbook_path: Path | str,
    *,
    sheet_name: str = "Wifi_LLAPI",
    column_overrides: dict[str, str] | None = None,
) -> dict[tuple[str, str], list[WorkbookRow]]:
    """Build a semantic-key workbook index.

    Multiple rows per semantic key are preserved as a list for later ambiguity
    handling.
    """

    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"empty sheet: {sheet_name}")

        if column_overrides:
            columns = {
                key: _column_letter_to_index(letter)
                for key, letter in column_overrides.items()
            }
            required = (
                "object",
                "api",
                "test_steps",
                "command_output",
                "result_5g",
                "result_6g",
                "result_24g",
            )
            missing = [key for key in required if key not in columns]
            if missing:
                raise ValueError(
                    f"column overrides missing required columns: {missing}"
                )
        else:
            columns = _auto_discover_columns(list(rows[0]))

        index: dict[tuple[str, str], list[WorkbookRow]] = {}
        for sheet_row_idx, row in enumerate(rows[1:], start=2):
            object_value = row[columns["object"]] if len(row) > columns["object"] else None
            api_value = row[columns["api"]] if len(row) > columns["api"] else None
            if not object_value or not api_value:
                continue

            key = (
                normalize_object(str(object_value)),
                normalize_api(str(api_value)),
            )
            workbook_row = WorkbookRow(
                raw_row_index=sheet_row_idx,
                object_path=str(object_value),
                api=str(api_value),
                test_steps=str(
                    row[columns["test_steps"]] if len(row) > columns["test_steps"] else ""
                ),
                command_output=str(
                    row[columns["command_output"]]
                    if len(row) > columns["command_output"]
                    else ""
                ),
                result_5g=str(
                    row[columns["result_5g"]] if len(row) > columns["result_5g"] else ""
                ),
                result_6g=str(
                    row[columns["result_6g"]] if len(row) > columns["result_6g"] else ""
                ),
                result_24g=str(
                    row[columns["result_24g"]] if len(row) > columns["result_24g"] else ""
                ),
            )
            index.setdefault(key, []).append(workbook_row)

        return index
    finally:
        workbook.close()
