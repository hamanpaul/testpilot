from __future__ import annotations

import shutil
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
for extra_path in (ROOT / 'src', ROOT / 'plugins' / 'wifi_llapi'):
    extra = str(extra_path)
    if extra not in sys.path:
        sys.path.insert(0, extra)

from openpyxl import Workbook, load_workbook
import pytest

from testpilot.core.orchestrator import Orchestrator
from testpilot.reporting.wifi_llapi_excel import ensure_template_report

FIXTURE_DIR = ROOT / 'tests' / 'fixtures' / 'wifi_llapi_delta'
ZERO_DELTA_COMMENT = 'fail 原因為 0，數值無變化'
CASE_IDS = [
    'wifi-llapi-delta-nonzero-pass',
    'wifi-llapi-delta-nonzero-fail',
    'wifi-llapi-delta-match-pass',
]
CASE_OUTPUTS = {
    'wifi-llapi-delta-nonzero-pass': {
        'ubus-cli "WiFi.SSID.4.Stats.BytesSent?"': 'BytesSent=100',
        'echo traffic-5g': 'traffic-5g',
        'ubus-cli "WiFi.SSID.4.Stats.BytesSentVerify?"': 'BytesSent=132',
    },
    'wifi-llapi-delta-nonzero-fail': {
        'ubus-cli "WiFi.SSID.6.Stats.BytesReceived?"': 'BytesReceived=88',
        'echo traffic-6g': 'traffic-6g',
        'ubus-cli "WiFi.SSID.6.Stats.BytesReceivedVerify?"': 'BytesReceived=88',
    },
    'wifi-llapi-delta-match-pass': {
        'ubus-cli "WiFi.SSID.8.Stats.PacketsSent?"': 'PacketsSent=10',
        "wl -i wl2 counters | grep '^pkt:'": 'DriverPacketsSent=20',
        'echo traffic-24g': 'traffic-24g',
        'ubus-cli "WiFi.SSID.8.Stats.PacketsSentVerify?"': 'PacketsSent=40',
        "wl -i wl2 counters | grep '^pkt:verify'": 'DriverPacketsSent=53',
    },
}


class MockTransport:
    def __init__(self, outputs: dict[str, str]) -> None:
        self.outputs = outputs
        self.history: list[str] = []
        self._connected = False

    @property
    def is_connected(self) -> bool:
        return self._connected

    def connect(self) -> None:
        self._connected = True

    def disconnect(self) -> None:
        self._connected = False

    def execute(self, command: str, timeout: float = 30.0) -> dict[str, Any]:
        self.history.append(command)
        return {
            'returncode': 0,
            'stdout': self.outputs.get(command, ''),
            'stderr': '',
            'elapsed': 0.01,
        }


def _write_testbed_yaml(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        '\n'.join(
            [
                'testbed:',
                '  name: delta-integration-testbed',
                '  serialwrap_binary: ./serialwrap-stub',
                '  devices:',
                '    DUT:',
                '      role: ap',
                '      transport: serial',
            ]
        )
        + '\n',
        encoding='utf-8',
    )


def _write_source_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wifi_LLAPI'
    ws['A2'] = 'Object'
    ws['C2'] = 'Parameter Name'
    ws['G2'] = 'Test steps'
    ws['H2'] = 'Command Output'
    ws['I2'] = 'ARC 4.0.3 Test Result\nWiFi 5g'
    ws['J2'] = 'ARC 4.0.3 Test Result\nWiFi 6g'
    ws['K2'] = 'ARC 4.0.3 Test Result\nWiFi 2.4g'
    ws['L2'] = 'Tester'
    ws['M2'] = 'ARC Comment'
    ws['A4'] = 'WiFi.SSID.{i}.Stats.'
    ws['C4'] = 'BytesSent'
    ws['A5'] = 'WiFi.SSID.{i}.Stats.'
    ws['C5'] = 'BytesReceived'
    ws['A6'] = 'WiFi.SSID.{i}.Stats.'
    ws['C6'] = 'PacketsSent'
    ws['A7'] = 'WiFi.SSID.{i}.Stats.'
    ws['C7'] = 'BlockedInvalid'
    wb.save(path)
    wb.close()


def _prepare_runtime_project(tmp_path: Path) -> Path:
    project_root = tmp_path / 'project'
    plugin_dir = project_root / 'plugins' / 'wifi_llapi'
    plugin_dir.mkdir(parents=True, exist_ok=True)

    source_plugin_dir = ROOT / 'plugins' / 'wifi_llapi'
    for name in (
        'plugin.py',
        'command_resolver.py',
        'baseline_qualifier.py',
        'agent-config.yaml',
        'band-baselines.yaml',
    ):
        shutil.copy2(source_plugin_dir / name, plugin_dir / name)

    _write_testbed_yaml(project_root / 'configs' / 'testbed.yaml')
    source_xlsx = project_root / 'source.xlsx'
    _write_source_xlsx(source_xlsx)
    ensure_template_report(
        source_xlsx=source_xlsx,
        template_path=plugin_dir / 'reports' / 'templates' / 'wifi_llapi_template.xlsx',
        manifest_path=plugin_dir / 'reports' / 'templates' / 'wifi_llapi_template.manifest.json',
    )
    return project_root


def _stage_cases(project_root: Path, names: list[str]) -> Path:
    cases_dir = project_root / 'plugins' / 'wifi_llapi' / 'cases'
    cases_dir.mkdir(parents=True, exist_ok=True)
    for name in names:
        shutil.copy2(FIXTURE_DIR / name, cases_dir / name)
    return cases_dir


def _patch_runtime(
    monkeypatch: pytest.MonkeyPatch,
    orch: Orchestrator,
) -> dict[str, MockTransport]:
    plugin = orch.loader.load('wifi_llapi')
    transports: dict[str, MockTransport] = {}

    def setup_env(case: dict[str, Any], topology: Any) -> bool:
        case_id = str(case.get('id', ''))
        transport = MockTransport(CASE_OUTPUTS[case_id])
        transport.connect()
        transports[case_id] = transport
        plugin._transports = {'DUT': transport}
        return True

    def verify_env(case: dict[str, Any], topology: Any) -> bool:
        return True

    def teardown(case: dict[str, Any], topology: Any) -> None:
        transport = plugin._transports.get('DUT')
        if transport is not None:
            transport.disconnect()
        plugin._transports.clear()

    monkeypatch.setattr(plugin, 'setup_env', setup_env)
    monkeypatch.setattr(plugin, 'verify_env', verify_env)
    monkeypatch.setattr(plugin, 'teardown', teardown)
    monkeypatch.setattr(orch, '_start_serialwrap_for_run', lambda: None)
    monkeypatch.setattr(
        orch,
        '_export_serialwrap_logs',
        lambda **kwargs: {},
    )
    monkeypatch.setattr(orch, '_stop_serialwrap', lambda: None)
    monkeypatch.setattr(orch.runner_selector, 'load_agent_config', lambda plugin_name: {})
    monkeypatch.setattr(
        orch.runner_selector,
        'build_execution_policy',
        lambda agent_config: {
            'retry': {'max_attempts': 1},
            'failure_policy': 'retry_then_fail_and_continue',
            'timeout': {
                'base_seconds': 1,
                'per_step_seconds': 0,
                'retry_multiplier': 1,
                'max_seconds': 1,
            },
        },
    )
    monkeypatch.setattr(
        orch.runner_selector,
        'select_case_runner',
        lambda plugin_name, case, agent_config: (
            {'provider': 'stub', 'model': 'test', 'reasoning_effort': 'low'},
            {'selected': {'provider': 'stub', 'model': 'test'}},
        ),
    )
    return transports


def _report_snapshot(report_path: Path) -> dict[int, dict[str, Any]]:
    wb = load_workbook(report_path)
    ws = wb['Wifi_LLAPI']
    snapshot = {
        row: {
            column: (ws[f'{column}{row}'].value or '')
            for column in ('G', 'H', 'I', 'J', 'K', 'M')
        }
        for row in (4, 5, 6)
    }
    wb.close()
    return snapshot


def test_delta_runtime_and_report_integration(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    project_root = _prepare_runtime_project(tmp_path)
    _stage_cases(
        project_root,
        [
            'delta_nonzero_pass.yaml',
            'delta_nonzero_fail.yaml',
            'delta_match_pass.yaml',
        ],
    )
    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / 'plugins',
        config_path=project_root / 'configs' / 'testbed.yaml',
    )
    transports = _patch_runtime(monkeypatch, orch)

    result = orch.run('wifi_llapi', case_ids=CASE_IDS, dut_fw_ver='FW-DELTA-IT-1')
    snapshot = _report_snapshot(Path(result['report_path']))

    assert result['cases_count'] == 3
    assert result['pass_count'] == 2
    assert result['fail_count'] == 1

    assert snapshot[4] == {
        'G': 'ubus-cli "WiFi.SSID.4.Stats.BytesSent?"\necho traffic-5g\nubus-cli "WiFi.SSID.4.Stats.BytesSentVerify?"',
        'H': 'BytesSent=100\ntraffic-5g\nBytesSent=132',
        'I': 'Pass',
        'J': 'N/A',
        'K': 'N/A',
        'M': '',
    }
    assert snapshot[5] == {
        'G': 'ubus-cli "WiFi.SSID.6.Stats.BytesReceived?"\necho traffic-6g\nubus-cli "WiFi.SSID.6.Stats.BytesReceivedVerify?"',
        'H': 'BytesReceived=88\ntraffic-6g\nBytesReceived=88',
        'I': 'N/A',
        'J': 'Fail',
        'K': 'N/A',
        'M': ZERO_DELTA_COMMENT,
    }
    assert snapshot[6] == {
        'G': (
            'ubus-cli "WiFi.SSID.8.Stats.PacketsSent?"\n'
            "wl -i wl2 counters | grep '^pkt:'\n"
            'echo traffic-24g\n'
            'ubus-cli "WiFi.SSID.8.Stats.PacketsSentVerify?"\n'
            "wl -i wl2 counters | grep '^pkt:verify'"
        ),
        'H': 'PacketsSent=10\nDriverPacketsSent=20\ntraffic-24g\nPacketsSent=40\nDriverPacketsSent=53',
        'I': 'N/A',
        'J': 'N/A',
        'K': 'Pass',
        'M': '',
    }
    assert transports['wifi-llapi-delta-nonzero-pass'].history == [
        'ubus-cli "WiFi.SSID.4.Stats.BytesSent?"',
        'echo traffic-5g',
        'ubus-cli "WiFi.SSID.4.Stats.BytesSentVerify?"',
    ]


def test_invalid_delta_schema_case_is_blocked_before_execution(tmp_path: Path) -> None:
    project_root = _prepare_runtime_project(tmp_path)
    template_path = project_root / 'plugins' / 'wifi_llapi' / 'reports' / 'templates' / 'wifi_llapi_template.xlsx'
    _stage_cases(
        project_root,
        [
            'delta_nonzero_pass.yaml',
            'delta_nonzero_fail.yaml',
            'delta_match_pass.yaml',
            'delta_invalid_blocked.yaml',
        ],
    )

    orch = Orchestrator(
        project_root=project_root,
        plugins_dir=project_root / 'plugins',
        config_path=project_root / 'configs' / 'testbed.yaml',
    )
    plugin = orch.loader.load('wifi_llapi')
    discovered_cases = sorted(plugin.discover_cases(), key=lambda case: str(case.get('id', '')))
    blocked_case = next(case for case in discovered_cases if case['id'] == 'wifi-llapi-delta-invalid-blocked')

    assert [case['id'] for case in discovered_cases] == [
        'wifi-llapi-delta-invalid-blocked',
        'wifi-llapi-delta-match-pass',
        'wifi-llapi-delta-nonzero-fail',
        'wifi-llapi-delta-nonzero-pass',
    ]
    assert blocked_case['llapi_support'] == 'Blocked'
    assert blocked_case['blocked_reason'] == (
        'invalid_delta_schema: delta_* operators require at least one phase=trigger step'
    )

    prep = orch._prepare_wifi_llapi_alignment(
        plugin=plugin,
        case_ids=CASE_IDS + ['wifi-llapi-delta-invalid-blocked'],
        template_path=template_path,
    )

    assert sorted(case['id'] for case in prep.runnable_cases) == sorted(CASE_IDS)
    assert [result.id_before for result in prep.blocked_results] == ['wifi-llapi-delta-invalid-blocked']
    assert prep.blocked_results[0].blocked_reason == (
        'invalid_delta_schema: delta_* operators require at least one phase=trigger step'
    )
