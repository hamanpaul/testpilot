from pathlib import Path
from collections import Counter

import pytest
import yaml
from openpyxl import Workbook, load_workbook

from testpilot.reporting.wifi_llapi_align import (
    AlignmentConflictError,
    AlignResult,
    TemplateIndex,
    _resolve_collisions,
    apply_alignment_mutations,
    align_case,
    build_template_index,
    write_blocked_cases_report,
    write_skipped_cases_report,
    _extract_name_api,
)

from testpilot.reporting.wifi_llapi_excel import create_run_report_from_template, fill_blocked_markers, fill_skip_markers
from testpilot.schema.case_schema import load_case

def test_extract_name_api_cases():
    # Should extract method token if present
    assert _extract_name_api("FailedRetransCount - WiFi.SSID.{i}.getSSIDStats().") == "getSSIDStats()"
    # Should extract left token if no method present
    assert _extract_name_api("AssociationTime - WiFi.AccessPoint.{i}.AssociatedDevice.{i}.") == "AssociationTime"
    # Regression: em-dash separator
    assert _extract_name_api("AssociationTime — WiFi.AccessPoint.{i}.AssociatedDevice.{i}.") == "AssociationTime"

def _build_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["A4"] = "WiFi.AccessPoint.{i}."
    ws["C4"] = "kickStation()"
    ws["A5"] = "WiFi.Radio.{i}."
    ws["C5"] = "getRadioStats()"
    ws["A6"] = "WiFi.AccessPoint.{i}.AssociatedDevice.{i}."
    ws["C6"] = "HeCapabilities"
    ws["A7"] = "WiFi.AccessPoint.{i}.AssociatedDevice.{i}."
    ws["C7"] = "DownlinkShortGuard"
    wb.save(path)
    wb.close()


@pytest.fixture
def template_path(tmp_path: Path) -> Path:
    template = tmp_path / "template.xlsx"
    _build_template(template)
    return template


def test_build_template_index_happy(template_path: Path):
    template = template_path

    index = build_template_index(template)

    assert index.forward[4] == ("WiFi.AccessPoint.{i}.", "kickStation()")
    assert index.by_object_api[("WiFi.Radio.{i}.", "getRadioStats()")] == 5
    assert index.by_api["HeCapabilities"] == [6]


def test_align_already_aligned(tmp_path: Path, template_path: Path):
    template = template_path
    index = build_template_index(template)
    case_file = tmp_path / "D006_hecapabilities.yaml"
    case_file.write_text("stub\n", encoding="utf-8")

    result = align_case(
        {
            "id": "wifi-llapi-D006-hecapabilities",
            "name": "HeCapabilities",
            "source": {
                "row": 6,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "HeCapabilities",
            },
        },
        index,
        case_file,
    )

    assert isinstance(result, AlignResult)
    assert result.status == "already_aligned"
    assert result.source_row_after == 6
    assert result.filename_after is None


def test_align_auto_source_row_drift(tmp_path: Path, template_path: Path):
    template = template_path
    index = build_template_index(template)
    case_file = tmp_path / "D021_hecapabilities.yaml"
    case_file.write_text("stub\n", encoding="utf-8")

    result = align_case(
        {
            "id": "wifi-llapi-D021-hecapabilities",
            "name": "HeCapabilities",
            "source": {
                "row": 7,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "HeCapabilities",
            },
        },
        index,
        case_file,
    )

    assert result.status == "auto_aligned"
    assert result.source_row_before == 7
    assert result.source_row_after == 6


def test_align_blocked_name_different_row(tmp_path: Path, template_path: Path):
    template = template_path
    index = build_template_index(template)
    case_file = tmp_path / "D021_hecapabilities.yaml"
    case_file.write_text("stub\n", encoding="utf-8")

    result = align_case(
        {
            "id": "wifi-llapi-D021-hecapabilities",
            "name": "DownlinkShortGuard",
            "source": {
                "row": 7,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "HeCapabilities",
            },
        },
        index,
        case_file,
    )

    assert result.status == "blocked"
    assert result.blocked_reason == "name_points_to_different_row"
    assert isinstance(index, TemplateIndex)


def test_align_skip_duplicate(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _build_template(template)
    index = build_template_index(template)
    winner = align_case(
        {
            "id": "wifi-llapi-D006-hecapabilities-a",
            "name": "HeCapabilities",
            "source": {
                "row": 6,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "HeCapabilities",
            },
        },
        index,
        tmp_path / "D006_hecapabilities_a.yaml",
    )
    loser = align_case(
        {
            "id": "wifi-llapi-D021-hecapabilities-b",
            "name": "HeCapabilities",
            "source": {
                "row": 21,
                "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                "api": "HeCapabilities",
            },
        },
        index,
        tmp_path / "D021_hecapabilities_b.yaml",
    )

    _resolve_collisions([winner, loser])

    assert winner.status in {"already_aligned", "auto_aligned"}
    assert loser.status == "skipped"
    assert loser.skip_winner_filename == "D006_hecapabilities_a.yaml"


def test_align_all_repo_cases():
    repo_root = Path(__file__).resolve().parents[3]
    template = repo_root / "plugins" / "wifi_llapi" / "reports" / "templates" / "wifi_llapi_template.xlsx"
    cases_dir = repo_root / "plugins" / "wifi_llapi" / "cases"
    index = build_template_index(template)
    case_paths = [
        path
        for path in sorted(cases_dir.glob("*.y*ml"))
        if not path.stem.startswith("_")
    ]
    cases = [(path, load_case(path)) for path in case_paths]
    results = [align_case(case, index, path) for path, case in cases]

    _resolve_collisions(results)

    case_ids = [str(case.get("id", "")).strip() for _path, case in cases]
    duplicates = sorted(case_id for case_id, count in Counter(case_ids).items() if count > 1)

    counts = {
        "already_aligned": sum(1 for r in results if r.status == "already_aligned"),
        "auto_aligned": sum(1 for r in results if r.status == "auto_aligned"),
        "blocked": sum(1 for r in results if r.status == "blocked"),
        "skipped": sum(1 for r in results if r.status == "skipped"),
    }

    assert sum(counts.values()) == 420
    assert counts["auto_aligned"] <= 168
    assert not duplicates, f"duplicate discoverable case ids: {duplicates}"


def test_apply_mutations_rename_collision(tmp_path: Path):
    source = tmp_path / "D021_hecapabilities.yaml"
    target = tmp_path / "D006_hecapabilities.yaml"
    source.write_text(
        yaml.safe_dump(
            {
                "id": "wifi-llapi-D021-hecapabilities",
                "name": "HeCapabilities",
                "source": {
                    "row": 21,
                    "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                    "api": "HeCapabilities",
                },
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    target.write_text("already here\n", encoding="utf-8")
    result = AlignResult(
        case_file=source,
        status="auto_aligned",
        source_row_before=21,
        source_row_after=6,
        source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        source_api="HeCapabilities",
        filename_before=source.name,
        filename_after=target.name,
        id_before="wifi-llapi-D021-hecapabilities",
        id_after="wifi-llapi-D006-hecapabilities",
        template_row=6,
        template_row_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        template_row_api="HeCapabilities",
    )

    with pytest.raises(AlignmentConflictError):
        apply_alignment_mutations([result])


def test_apply_mutations_rewrite_rename_and_update_path(tmp_path: Path):
    source = tmp_path / "D021_hecapabilities.yaml"
    destination = tmp_path / "D006_hecapabilities.yaml"
    source.write_text(
        yaml.safe_dump(
            {
                "id": "wifi-llapi-D021-hecapabilities",
                "name": "HeCapabilities",
                "source": {
                    "row": 21,
                    "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                    "api": "HeCapabilities",
                },
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    result = AlignResult(
        case_file=source,
        status="auto_aligned",
        source_row_before=21,
        source_row_after=6,
        source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        source_api="HeCapabilities",
        filename_before=source.name,
        filename_after=destination.name,
        id_before="wifi-llapi-D021-hecapabilities",
        id_after="wifi-llapi-D006-hecapabilities",
        template_row=6,
        template_row_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        template_row_api="HeCapabilities",
    )

    apply_alignment_mutations([result])

    assert not source.exists()
    assert destination.exists()
    payload = yaml.safe_load(destination.read_text(encoding="utf-8"))
    assert payload["source"]["row"] == 6
    assert payload["id"] == "wifi-llapi-D006-hecapabilities"
    assert result.case_file == destination


def test_apply_mutations_restore_source_when_second_replace_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    source = tmp_path / "D021_hecapabilities.yaml"
    destination = tmp_path / "D006_hecapabilities.yaml"
    source.write_text(
        yaml.safe_dump(
            {
                "id": "wifi-llapi-D021-hecapabilities",
                "name": "HeCapabilities",
                "source": {
                    "row": 21,
                    "object": "WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
                    "api": "HeCapabilities",
                },
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    result = AlignResult(
        case_file=source,
        status="auto_aligned",
        source_row_before=21,
        source_row_after=6,
        source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        source_api="HeCapabilities",
        filename_before=source.name,
        filename_after=destination.name,
        id_before="wifi-llapi-D021-hecapabilities",
        id_after="wifi-llapi-D006-hecapabilities",
        template_row=6,
        template_row_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
        template_row_api="HeCapabilities",
    )

    original_replace = Path.replace
    replace_calls: list[tuple[str, str]] = []

    def _replace_with_second_call_failure(self: Path, target: Path | str) -> Path:
        target_path = Path(target)
        replace_calls.append((self.name, target_path.name))
        if len(replace_calls) == 2:
            raise OSError(f"boom: {self} -> {target_path}")
        return original_replace(self, target_path)

    monkeypatch.setattr(Path, "replace", _replace_with_second_call_failure)

    with pytest.raises(OSError, match="boom"):
        apply_alignment_mutations([result])

    payload = yaml.safe_load(source.read_text(encoding="utf-8"))
    assert payload["source"]["row"] == 21
    assert payload["id"] == "wifi-llapi-D021-hecapabilities"
    assert not destination.exists()
    assert sorted(path.name for path in tmp_path.iterdir()) == [source.name]
    assert result.case_file == source
    assert len(replace_calls) == 3
    assert replace_calls[0][0] == source.name
    assert replace_calls[0][1].startswith(f".{source.name}.")
    assert replace_calls[0][1].endswith(".bak")
    assert replace_calls[1][0].startswith(f".{destination.name}.")
    assert replace_calls[1][0].endswith(".tmp")
    assert replace_calls[1][1] == destination.name
    assert replace_calls[2][0] == replace_calls[0][1]
    assert replace_calls[2][1] == source.name


def test_write_blocked_report_md(tmp_path: Path):
    out_path = tmp_path / "blocked_cases.md"
    blocked = [
        AlignResult(
            case_file=tmp_path / "D021_hecapabilities.yaml",
            status="blocked",
            source_row_before=18,
            source_row_after=None,
            source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            source_api="HeCapabilities",
            filename_before="D021_hecapabilities.yaml",
            filename_after=None,
            id_before="wifi-llapi-D021-hecapabilities",
            id_after=None,
            blocked_reason="name_points_to_different_row",
            template_row=21,
            template_row_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            template_row_api="DownlinkShortGuard",
        )
    ]

    write_blocked_cases_report(blocked, out_path)

    text = out_path.read_text(encoding="utf-8")
    assert "| case_id | filename | source.row |" in text
    assert "name_points_to_different_row" in text


def test_write_skipped_report_md(tmp_path: Path):
    out_path = tmp_path / "skipped_cases.md"
    skipped = [
        AlignResult(
            case_file=tmp_path / "D030_dup.yaml",
            status="skipped",
            source_row_before=21,
            source_row_after=21,
            source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            source_api="HeCapabilities",
            filename_before="D030_dup.yaml",
            filename_after=None,
            id_before="wifi-llapi-D030-dup",
            id_after=None,
            skip_winner_filename="D021_hecapabilities.yaml",
            template_row=21,
            template_row_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            template_row_api="HeCapabilities",
        )
    ]

    write_skipped_cases_report(skipped, out_path)

    text = out_path.read_text(encoding="utf-8")
    assert "| case_id | filename | source.row | winner_filename | template_N |" in text
    assert "D021_hecapabilities.yaml" in text


def test_fill_blocked_markers(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _build_template(template)
    report = tmp_path / "report.xlsx"
    create_run_report_from_template(template, report)

    blocked = [
        AlignResult(
            case_file=tmp_path / "D021_hecapabilities.yaml",
            status="blocked",
            source_row_before=6,
            source_row_after=None,
            source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            source_api="HeCapabilities",
            filename_before="D021_hecapabilities.yaml",
            filename_after=None,
            id_before="wifi-llapi-D021-hecapabilities",
            id_after=None,
            blocked_reason="name_not_in_template",
        )
    ]

    fill_blocked_markers(report, blocked)
    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["H6"].value == "BLOCKED: name_not_in_template"
    assert ws["G6"].value is None
    assert ws["I6"].value is None
    wb.close()


def test_fill_skip_markers(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _build_template(template)
    report = tmp_path / "report.xlsx"
    create_run_report_from_template(template, report)

    skipped = [
        AlignResult(
            case_file=tmp_path / "D030_dup.yaml",
            status="skipped",
            source_row_before=6,
            source_row_after=6,
            source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            source_api="HeCapabilities",
            filename_before="D030_dup.yaml",
            filename_after=None,
            id_before="wifi-llapi-D030-dup",
            id_after=None,
            skip_winner_filename="D006_hecapabilities.yaml",
            template_row=6,
        )
    ]

    fill_skip_markers(report, skipped)
    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    assert ws["H6"].value == "SKIP: duplicate with D006"
    assert ws["G6"].value is None
    assert ws["I6"].value is None
    wb.close()


def test_fill_skip_markers_out_of_range(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _build_template(template)
    report = tmp_path / "report.xlsx"
    create_run_report_from_template(template, report)

    # Use a template_row that is out of worksheet bounds
    skipped = [
        AlignResult(
            case_file=tmp_path / "D031_dup.yaml",
            status="skipped",
            source_row_before=6,
            source_row_after=6,
            source_object="WiFi.AccessPoint.{i}.AssociatedDevice.{i}.",
            source_api="HeCapabilities",
            filename_before="D031_dup.yaml",
            filename_after=None,
            id_before="wifi-llapi-D031-dup",
            id_after=None,
            skip_winner_filename="D021_hecapabilities.yaml",
            template_row=999,
        )
    ]

    fill_skip_markers(report, skipped)
    wb = load_workbook(report)
    ws = wb["Wifi_LLAPI"]
    # Out-of-range template_row should cause the marker to be skipped
    assert ws["H6"].value is None
    assert ws["G6"].value is None
    assert ws["I6"].value is None
    wb.close()
