from __future__ import annotations

from dataclasses import dataclass
import logging
from pathlib import Path
import re
from typing import Literal
from uuid import uuid4

from openpyxl import load_workbook
import yaml

from testpilot.reporting.wifi_llapi_excel import (
    DATA_START_ROW,
    DEFAULT_SHEET_NAME,
    EMPTY_STREAK_STOP,
    MAX_SCAN_ROWS,
    normalize_text,
)

AlignStatus = Literal["already_aligned", "auto_aligned", "blocked", "skipped"]
BlockedReason = Literal[
    "object_api_not_in_template",
    "name_points_to_different_row",
    "name_not_in_template",
    "ambiguous_object_api_family",
]

_ID_D_PATTERN = re.compile(r"(wifi-llapi-D)(\d{3})(-)")
_FILE_D_PATTERN = re.compile(r"^D(?P<row>\d{3})(?P<suffix>.*)$")

log = logging.getLogger(__name__)


@dataclass(slots=True)
class TemplateIndex:
    forward: dict[int, tuple[str, str]]
    by_object_api: dict[tuple[str, str], list[int]]
    by_api: dict[str, list[int]]


@dataclass(slots=True)
class AlignResult:
    case_file: Path
    status: AlignStatus
    source_row_before: int
    source_row_after: int | None
    source_object: str
    source_api: str
    filename_before: str
    filename_after: str | None
    id_before: str
    id_after: str | None
    blocked_reason: BlockedReason | None = None
    skip_winner_filename: str | None = None
    template_row: int | None = None
    template_row_object: str | None = None
    template_row_api: str | None = None
    candidate_template_rows: list[int] | None = None


class AlignmentConflictError(RuntimeError):
    pass


_METHOD_TOKEN_PATTERN = re.compile(r"\b[A-Za-z_][A-Za-z0-9_]*\(\)")

def _extract_name_api(name: object) -> str:
    text = str(name).strip() if isinstance(name, str) else ""
    if not text:
        return ""
    method_tokens = _METHOD_TOKEN_PATTERN.findall(text)
    if method_tokens:
        return method_tokens[-1]
    for separator in (" - ", " — "):
        if separator in text:
            return text.split(separator, 1)[0].strip()
    if "." in text:
        return text.rsplit(".", 1)[-1].strip()
    return text


def _replace_id_row(case_id: str, canonical_row: int) -> str | None:
    if not _ID_D_PATTERN.search(case_id):
        return None
    return _ID_D_PATTERN.sub(rf"\g<1>{canonical_row:03d}\g<3>", case_id, count=1)


def build_template_index(template_xlsx: Path) -> TemplateIndex:
    wb = load_workbook(template_xlsx, read_only=True, data_only=True)
    try:
        ws = wb[DEFAULT_SHEET_NAME]
        forward: dict[int, tuple[str, str]] = {}
        by_object_api: dict[tuple[str, str], list[int]] = {}
        by_api: dict[str, list[int]] = {}
        empty_streak = 0
        for row in range(DATA_START_ROW, min(ws.max_row, DATA_START_ROW + MAX_SCAN_ROWS) + 1):
            obj = normalize_text(ws[f"A{row}"].value)
            api = normalize_text(ws[f"C{row}"].value)
            if not api:
                empty_streak += 1
                if empty_streak >= EMPTY_STREAK_STOP:
                    break
                continue
            empty_streak = 0
            forward[row] = (obj, api)
            by_object_api.setdefault((obj, api), []).append(row)
            by_api.setdefault(api, []).append(row)
        return TemplateIndex(forward=forward, by_object_api=by_object_api, by_api=by_api)
    finally:
        wb.close()


def align_case(case: dict, index: TemplateIndex, case_file: Path) -> AlignResult:
    source = case.get("source") if isinstance(case.get("source"), dict) else {}
    obj = normalize_text(source.get("object"))
    api = normalize_text(source.get("api"))
    source_row_before = int(source.get("row", 0) or 0)
    case_id = str(case.get("id", ""))
    name_api = _extract_name_api(case.get("name"))
    filename_before = case_file.name
    candidate_rows = index.by_object_api.get((obj, api), [])
    if not candidate_rows:
        return AlignResult(
            case_file=case_file,
            status="blocked",
            source_row_before=source_row_before,
            source_row_after=None,
            source_object=obj,
            source_api=api,
            filename_before=filename_before,
            filename_after=None,
            id_before=case_id,
            id_after=None,
            blocked_reason="object_api_not_in_template",
        )
    template_row: int | None = None
    if len(candidate_rows) > 1:
        if source_row_before in candidate_rows:
            template_row = source_row_before
        else:
            return AlignResult(
                case_file=case_file,
                status="blocked",
                source_row_before=source_row_before,
                source_row_after=None,
                source_object=obj,
                source_api=api,
                filename_before=filename_before,
                filename_after=None,
                id_before=case_id,
                id_after=None,
                blocked_reason="ambiguous_object_api_family",
                candidate_template_rows=list(candidate_rows),
            )
    else:
        template_row = candidate_rows[0]
    template_object, template_api = index.forward.get(template_row, ("", ""))
    if name_api and name_api != template_api:
        name_api_candidates = index.by_api.get(name_api, [])
        reason = "name_points_to_different_row" if name_api_candidates else "name_not_in_template"
        return AlignResult(
            case_file=case_file,
            status="blocked",
            source_row_before=source_row_before,
            source_row_after=None,
            source_object=obj,
            source_api=api,
            filename_before=filename_before,
            filename_after=None,
            id_before=case_id,
            id_after=None,
            blocked_reason=reason,
            template_row=template_row,
            template_row_object=template_object,
            template_row_api=template_api,
            candidate_template_rows=list(name_api_candidates) if name_api_candidates else None,
        )
    filename_after = None
    source_row_after = template_row
    id_after = _replace_id_row(case_id, template_row)
    if not filename_before.startswith(f"D{template_row:03d}_"):
        match = _FILE_D_PATTERN.match(case_file.stem)
        suffix = match.group("suffix") if match else f"_{case_file.stem}"
        filename_after = f"D{template_row:03d}{suffix}{case_file.suffix}"
    if source_row_before == template_row and filename_after is None and id_after in (None, case_id):
        return AlignResult(
            case_file=case_file,
            status="already_aligned",
            source_row_before=source_row_before,
            source_row_after=template_row,
            source_object=obj,
            source_api=api,
            filename_before=filename_before,
            filename_after=None,
            id_before=case_id,
            id_after=None,
            template_row=template_row,
            template_row_object=template_object,
            template_row_api=template_api,
        )
    return AlignResult(
        case_file=case_file,
        status="auto_aligned",
        source_row_before=source_row_before,
        source_row_after=template_row,
        source_object=obj,
        source_api=api,
        filename_before=filename_before,
        filename_after=filename_after,
        id_before=case_id,
        id_after=id_after if id_after != case_id else None,
        template_row=template_row,
        template_row_object=template_object,
        template_row_api=template_api,
    )


def _resolve_collisions(results: list[AlignResult]) -> None:
    winners: dict[int, str] = {}
    runnable = sorted(
        [r for r in results if r.status in {"already_aligned", "auto_aligned"} and r.template_row is not None],
        key=lambda item: item.filename_before,
    )
    for result in runnable:
        if result.template_row is None:
            continue
        winner = winners.get(result.template_row)
        if winner is None:
            winners[result.template_row] = result.filename_before
            continue
        result.status = "skipped"
        result.skip_winner_filename = winner
        result.filename_after = None
        result.id_after = None


def apply_alignment_mutations(results: list[AlignResult]) -> None:
    for result in results:
        if result.status != "auto_aligned":
            continue
        source_path = result.case_file
        destination = source_path
        if result.filename_after:
            destination = source_path.with_name(result.filename_after)
            if destination.exists() and destination != source_path:
                raise AlignmentConflictError(f"alignment rename target already exists: {destination}")
        with source_path.open("r", encoding="utf-8") as handle:
            payload = yaml.safe_load(handle) or {}
        payload.setdefault("source", {})
        payload["source"]["row"] = result.source_row_after
        if result.id_after:
            payload["id"] = result.id_after
        encoded = yaml.safe_dump(payload, sort_keys=False, allow_unicode=True)
        temp_destination = destination.with_name(f".{destination.name}.{uuid4().hex}.tmp")
        backup_source: Path | None = None
        try:
            temp_destination.write_text(encoded, encoding="utf-8")
            if destination == source_path:
                temp_destination.replace(destination)
            else:
                backup_source = source_path.with_name(f".{source_path.name}.{uuid4().hex}.bak")
                source_path.replace(backup_source)
                try:
                    temp_destination.replace(destination)
                except Exception:
                    if backup_source.exists():
                        backup_source.replace(source_path)
                        backup_source = None
                    raise
                if backup_source.exists():
                    backup_source.unlink()
                    backup_source = None
            result.case_file = destination
        finally:
            if temp_destination.exists():
                temp_destination.unlink()
            if backup_source is not None and backup_source.exists():
                if source_path.exists():
                    backup_source.unlink()
                else:
                    backup_source.replace(source_path)
        log.info(
            "aligned: %s -> %s (source.row %s -> %s)",
            result.filename_before,
            result.case_file.name,
            result.source_row_before,
            result.source_row_after,
        )


def write_blocked_cases_report(blocked: list[AlignResult], out_path: Path) -> None:
    if not blocked:
        return
    lines = [
        "| case_id | filename | source.row | source.(object, api) | template_row_(object, api) | reason |",
        "|---|---|---:|---|---|---|",
    ]
    for item in blocked:
        source_pair = f"({item.source_object or '—'}, {item.source_api or '—'})"
        template_pair = (
            f"({item.template_row_object}, {item.template_row_api})"
            if item.template_row_object and item.template_row_api
            else "—"
        )
        if item.candidate_template_rows:
            template_pair = f"{template_pair} @ rows {item.candidate_template_rows}"
        lines.append(
            f"| {item.id_before} | {item.filename_before} | {item.source_row_before} | "
            f"{source_pair} | {template_pair} | {item.blocked_reason} |"
        )
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_skipped_cases_report(skipped: list[AlignResult], out_path: Path) -> None:
    if not skipped:
        return
    lines = [
        "| case_id | filename | source.row | winner_filename | template_N |",
        "|---|---|---:|---|---:|",
    ]
    for item in skipped:
        lines.append(
            f"| {item.id_before} | {item.filename_before} | {item.source_row_before} | "
            f"{item.skip_winner_filename or '—'} | {item.template_row if item.template_row is not None else '—'} |"
        )
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
