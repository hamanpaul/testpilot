from __future__ import annotations

import subprocess
import sys
from pathlib import Path
import pytest

import yaml
from openpyxl import Workbook

from testpilot.reporting.wifi_llapi_inventory import (
    audit_wifi_llapi_inventory,
    build_wifi_llapi_inventory_reconcile_plan,
    apply_wifi_llapi_inventory_reconcile_plan,
)


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wifi_LLAPI"
    ws["A4"] = "WiFi.Radio.{i}."
    ws["C4"] = "getRadioStats()"
    ws["A5"] = "WiFi.SSID.{i}."
    ws["C5"] = "getSSIDStats()"
    wb.save(path)
    wb.close()


def _write_case(path: Path, *, case_id: str, row: int, obj: str, api: str) -> None:
    payload = {
        "id": case_id,
        "name": api,
        "topology": {"devices": {"DUT": {}, "STA": {}}},
        "steps": [{"id": "step-1", "action": "noop", "target": "DUT"}],
        "pass_criteria": [{"field": "status", "operator": "equals", "expected": "Pass"}],
        "source": {"row": row, "object": obj, "api": api},
    }
    path.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")


def _init_repo(repo_root: Path) -> None:
    subprocess.run(["git", "init"], cwd=repo_root, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "test@example.com"], cwd=repo_root, check=True)
    subprocess.run(["git", "config", "user.name", "Test User"], cwd=repo_root, check=True)


def test_reconcile_plan_reports_restore_rewrite_and_demote_actions(tmp_path: Path) -> None:
    repo_root = tmp_path / "repo"
    cases_dir = repo_root / "plugins" / "wifi_llapi" / "cases"
    templates_dir = repo_root / "plugins" / "wifi_llapi" / "reports" / "templates"
    cases_dir.mkdir(parents=True)
    templates_dir.mkdir(parents=True)
    _write_template(templates_dir / "wifi_llapi_template.xlsx")

    _init_repo(repo_root)

    restored_case = cases_dir / "D004_restored.yaml"
    _write_case(restored_case, case_id="wifi-llapi-D004-restored", row=4, obj="WiFi.Radio.{i}.", api="getRadioStats()")
    subprocess.run(["git", "add", "."], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "seed history"], cwd=repo_root, check=True, capture_output=True)
    restored_case.unlink()
    subprocess.run(["git", "add", "-u"], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "remove canonical row"], cwd=repo_root, check=True, capture_output=True)

    _write_case(
        cases_dir / "D999_drifted.yaml",
        case_id="wifi-llapi-D999-drifted",
        row=999,
        obj="WiFi.SSID.{i}.",
        api="getSSIDStats()",
    )
    _write_case(
        cases_dir / "D777_extra.yaml",
        case_id="wifi-llapi-D777-extra",
        row=777,
        obj="WiFi.Unknown.{i}.",
        api="notInTemplate()",
    )

    plan = build_wifi_llapi_inventory_reconcile_plan(
        templates_dir / "wifi_llapi_template.xlsx",
        cases_dir,
        repo_root=repo_root,
    )

    assert any(action.kind == "restore" and action.row == 4 for action in plan.actions)
    assert any(action.kind == "rewrite" and action.row == 5 for action in plan.actions)
    assert any(action.kind == "demote" for action in plan.actions)
    assert not plan.blockers

    dry_run = plan.to_lines()
    assert any(line.startswith("restore ") for line in dry_run)
    assert any(line.startswith("rewrite ") for line in dry_run)
    assert any(line.startswith("demote ") for line in dry_run)

    script = Path(__file__).resolve().parents[3] / "scripts" / "wifi_llapi_reconcile_inventory.py"
    result = subprocess.run(
        [
            sys.executable,
            str(script),
            "--repo-root",
            str(repo_root),
            "--template-xlsx",
            str(templates_dir / "wifi_llapi_template.xlsx"),
            "--cases-dir",
            str(cases_dir),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    assert "dry-run:" in result.stdout
    assert "restore " in result.stdout
    assert "rewrite " in result.stdout
    assert "demote " in result.stdout

    apply_wifi_llapi_inventory_reconcile_plan(plan)
    audit = audit_wifi_llapi_inventory(templates_dir / "wifi_llapi_template.xlsx", cases_dir)
    assert audit.missing_rows == ()
    assert audit.case_status_counts["extra"] == 0
    assert audit.rows[4].status == "canonical"
    assert audit.rows[5].status == "canonical"


def test_ambiguous_history_candidate_becomes_blocker(tmp_path: Path) -> None:
    repo_root = tmp_path / "repo"
    cases_dir = repo_root / "plugins" / "wifi_llapi" / "cases"
    templates_dir = repo_root / "plugins" / "wifi_llapi" / "reports" / "templates"
    cases_dir.mkdir(parents=True)
    templates_dir.mkdir(parents=True)
    _write_template(templates_dir / "wifi_llapi_template.xlsx")
    _init_repo(repo_root)

    _write_case(
        cases_dir / "D004_alpha.yaml",
        case_id="wifi-llapi-D004-alpha",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )
    subprocess.run(["git", "add", "."], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "alpha"], cwd=repo_root, check=True, capture_output=True)
    (cases_dir / "D004_alpha.yaml").unlink()
    _write_case(
        cases_dir / "D004_beta.yaml",
        case_id="wifi-llapi-D004-beta",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )
    subprocess.run(["git", "add", "."], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "beta"], cwd=repo_root, check=True, capture_output=True)
    (cases_dir / "D004_beta.yaml").unlink()
    subprocess.run(["git", "add", "-u"], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "remove row"], cwd=repo_root, check=True, capture_output=True)

    plan = build_wifi_llapi_inventory_reconcile_plan(
        templates_dir / "wifi_llapi_template.xlsx",
        cases_dir,
        repo_root=repo_root,
    )

    assert any(action.reason == "ambiguous_history_candidate" for action in plan.blockers)
    assert not any(action.kind == "restore" and action.row == 4 for action in plan.actions)
    with pytest.raises(ValueError):
        apply_wifi_llapi_inventory_reconcile_plan(plan)


def test_restore_target_occupied_by_demoted_file_is_applied_safely(tmp_path: Path) -> None:
    repo_root = tmp_path / "repo"
    cases_dir = repo_root / "plugins" / "wifi_llapi" / "cases"
    templates_dir = repo_root / "plugins" / "wifi_llapi" / "reports" / "templates"
    cases_dir.mkdir(parents=True)
    templates_dir.mkdir(parents=True)
    _write_template(templates_dir / "wifi_llapi_template.xlsx")
    _init_repo(repo_root)

    _write_case(
        cases_dir / "D004_restoretarget.yaml",
        case_id="wifi-llapi-D004-restoretarget",
        row=4,
        obj="WiFi.Radio.{i}.",
        api="getRadioStats()",
    )
    subprocess.run(["git", "add", "."], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "seed history"], cwd=repo_root, check=True, capture_output=True)
    _write_case(
        cases_dir / "D004_restoretarget.yaml",
        case_id="wifi-llapi-D004-restoretarget",
        row=999,
        obj="WiFi.Unknown.{i}.",
        api="notInTemplate()",
    )
    subprocess.run(["git", "add", "."], cwd=repo_root, check=True)
    subprocess.run(["git", "commit", "-m", "drift current file"], cwd=repo_root, check=True, capture_output=True)
    _write_case(
        cases_dir / "D005_canonical.yaml",
        case_id="wifi-llapi-D005-canonical",
        row=5,
        obj="WiFi.SSID.{i}.",
        api="getSSIDStats()",
    )

    plan = build_wifi_llapi_inventory_reconcile_plan(
        templates_dir / "wifi_llapi_template.xlsx",
        cases_dir,
        repo_root=repo_root,
    )

    assert any(action.kind == "restore" and action.row == 4 for action in plan.actions)
    assert any(action.kind == "demote" and action.case_file.name == "D004_restoretarget.yaml" for action in plan.actions)
    apply_wifi_llapi_inventory_reconcile_plan(plan)
    audit = audit_wifi_llapi_inventory(templates_dir / "wifi_llapi_template.xlsx", cases_dir)
    assert audit.rows[4].status == "canonical"
    assert audit.rows[5].status == "canonical"
    assert all(len(a.discoverable_case_files) == 1 for a in audit.rows.values())


def test_repo_scale_inventory_reports_drifted_cases_without_omission() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    audit = audit_wifi_llapi_inventory(
        repo_root / "plugins" / "wifi_llapi" / "reports" / "templates" / "wifi_llapi_template.xlsx",
        repo_root / "plugins" / "wifi_llapi" / "cases",
    )
    plan = build_wifi_llapi_inventory_reconcile_plan(
        repo_root / "plugins" / "wifi_llapi" / "reports" / "templates" / "wifi_llapi_template.xlsx",
        repo_root / "plugins" / "wifi_llapi" / "cases",
        repo_root=repo_root,
    )

    omitted = {"D068_discoverymethodenabled_accesspoint_rnr.yaml", "D495_retrycount_ssid_stats_verified.yaml"}
    present = {action.case_file.name for action in plan.actions if action.case_file is not None}
    present |= {action.case_file.name for action in plan.blockers if action.case_file is not None}
    assert omitted <= present
